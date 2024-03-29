#!/usr/bin/python3
from bilibili_api.live import LiveDanmaku, LiveRoom
import sys
from openpyxl import Workbook, load_workbook
import json
import time
import os
import asyncio

async def get_room_play_info(roomid):
    liveRoom = LiveRoom(room_display_id=roomid)
    return await liveRoom.get_room_play_info()
async def get_room_info(roomid):
    liveRoom = LiveRoom(room_display_id=roomid)
    return await liveRoom.get_room_info()

class DMK:
    def __init__(self, roomid):
        self.roomid = roomid

    async def mkexcel(self):
        self.realid = (await get_room_play_info(self.roomid))['room_id']
        self.title = (await get_room_info(self.roomid))['room_info']['title']
        if '{}-{}.xlsx'.format(self.title, time.strftime('%Y-%m-%d', time.localtime())) in os.listdir('{}/out'.format(os.path.dirname(os.path.realpath(__file__)))):
            self.excel = load_workbook('out/{}-{}.xlsx'.format(self.title, time.strftime('%Y-%m-%d', time.localtime())))
            sheet = self.excel.active
        else:
            self.excel = Workbook()
            sheet = self.excel.active
            sheet['A1'] = 'user'
            sheet['B1'] = 'uid'
            sheet['C1'] = 'timestamp'
            sheet['D1'] = 'content'

        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 40

    def push(self, danmaku):
        sheet = self.excel.active
        sheet.append(danmaku)

    def close(self):
        self.excel.save(
            'out/{}-{}.xlsx'.format(
                self.title, 
                time.strftime('%Y-%m-%d', time.localtime())
            )
        )

dmk = None
room = None

async def main():
    global room
    global dmk
    
    try:
        roomid = int(sys.argv[1])
    except Exception:
        print("请输入直播间号码")
        return

    room = LiveDanmaku(room_display_id=(await get_room_play_info(roomid))['room_id'])
    dmk = DMK(roomid)

    async def on_danmu(msg):
        user = msg['data']['info'][2][1]
        uid = msg['data']['info'][2][0]
        timestamp = time.strftime(
            '%Y-%m-%d %H:%M:%S',
            time.localtime(msg['data']['info'][0][4]/1000)
        )
        content = msg['data']['info'][1]
        print(
            '\033[34m[{timestamp}]\033[0m\033[36m[{user}]\033[0m\033[35m[{uid}]\033[0m {content}'
            .format(
                user=user,
                uid=uid,
                timestamp=timestamp,
                content=content
            )
        )
        dmk.push([user, uid, timestamp, content])

    room.add_event_listener('DANMU_MSG', on_danmu)

    await dmk.mkexcel()
    await room.connect()

try:
    asyncio.get_event_loop().run_until_complete(main())
except KeyboardInterrupt:
    pass
except Exception as e:
    print(e)
finally:
    dmk.close()
    dmk = None